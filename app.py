import os
import sqlite3
import json
import glob
import csv
import io
import traceback
from datetime import datetime
from flask import Flask, request, jsonify, redirect

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE_DIR, "db.sqlite3")

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
SCHEMA = """
CREATE TABLE IF NOT EXISTS core_product (
    id INTEGER PRIMARY KEY,
    product_number VARCHAR(100),
    product_name TEXT,
    product_description TEXT,
    image_url VARCHAR(500),
    all_image_urls TEXT,
    source_category VARCHAR(255),
    source_subcategory VARCHAR(255),
    collection_name VARCHAR(255),
    materials TEXT,
    product_dimensions TEXT,
    product_weight REAL,
    country_of_origin VARCHAR(100),
    product_url VARCHAR(500),
    model_number VARCHAR(100),
    bullets TEXT,
    brand VARCHAR(255),
    product_type VARCHAR(255)
);
CREATE TABLE IF NOT EXISTS core_taxonomynode (
    id INTEGER PRIMARY KEY,
    shopify_id VARCHAR(64) UNIQUE,
    name VARCHAR(255),
    level SMALLINT,
    keywords TEXT,
    product_type_hint VARCHAR(255),
    parent_id BIGINT,
    attributes TEXT
);
CREATE TABLE IF NOT EXISTS core_classification (
    id INTEGER PRIMARY KEY,
    confidence REAL,
    status VARCHAR(20),
    alternatives TEXT,
    detected_attributes TEXT,
    needs_review BOOL,
    review_reason VARCHAR(255),
    classified_at DATETIME,
    reviewed_at DATETIME,
    reviewer_notes TEXT,
    product_id BIGINT,
    taxonomy_node_id BIGINT
);
CREATE TABLE IF NOT EXISTS core_batch (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name VARCHAR(255),
    status VARCHAR(20),
    total_products INTEGER,
    processed INTEGER DEFAULT 0,
    succeeded INTEGER DEFAULT 0,
    failed INTEGER DEFAULT 0,
    started_at DATETIME,
    finished_at DATETIME,
    error_log TEXT,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_class_product ON core_classification(product_id);
CREATE INDEX IF NOT EXISTS idx_class_status ON core_classification(status);
CREATE INDEX IF NOT EXISTS idx_class_node ON core_classification(taxonomy_node_id);
CREATE INDEX IF NOT EXISTS idx_product_number ON core_product(product_number);
"""

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------
def get_db():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA journal_mode=WAL")
    return c

def init_db():
    need = False
    if not os.path.exists(DB):
        need = True
    else:
        try:
            c = sqlite3.connect(DB)
            count = c.execute("SELECT COUNT(*) FROM core_product").fetchone()[0]
            c.close()
            if count == 0:
                need = True
        except:
            need = True
    if not need:
        return
    print("Seeding database...")
    c = sqlite3.connect(DB)
    c.executescript(SCHEMA)
    chunks = sorted(glob.glob(os.path.join(BASE_DIR, "seed_chunk_*.json")))
    if not chunks:
        c.close()
        return
    prods, nodes, cls = [], None, []
    for p in chunks:
        try:
            d = json.load(open(p, encoding="utf-8"))
            prods.extend(d.get("products", []))
            cls.extend(d.get("classifications", []))
            if nodes is None:
                nodes = d.get("taxonomy_nodes")
        except Exception as e:
            print(f"  Error loading {p}: {e}")
    cur = c.cursor()
    if nodes:
        for n in nodes:
            attrs = json.dumps(n.get("attributes", []))
            cur.execute("INSERT OR IGNORE INTO core_taxonomynode VALUES (?,?,?,?,?,?,?,?)",
                (n["id"], n["shopify_id"], n["name"], n["level"],
                 n.get("keywords", ""), n.get("product_type_hint", ""),
                 n.get("parent_id"), attrs))
    for p in prods:
        try:
            cur.execute("INSERT OR IGNORE INTO core_product VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (p["id"], p.get("product_number"), p.get("product_name"),
                 p.get("product_description"), p.get("image_url"),
                 p.get("all_image_urls"), p.get("source_category"),
                 p.get("source_subcategory"), p.get("collection_name"),
                 p.get("materials"), p.get("product_dimensions"),
                 p.get("product_weight"), p.get("country_of_origin"),
                 p.get("product_url"), p.get("model_number"),
                 p.get("bullets"), p.get("brand", ""), p.get("product_type", "")))
        except Exception as e:
            print(f"  Error inserting product {p.get('product_number')}: {e}")
    for cl in cls:
        try:
            cur.execute("INSERT OR IGNORE INTO core_classification VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (cl["id"], cl["confidence"], cl["status"],
                 cl.get("alternatives"), cl.get("detected_attributes"),
                 cl["needs_review"], cl.get("review_reason"),
                 cl.get("classified_at"), cl.get("reviewed_at"),
                 cl.get("reviewer_notes"), cl["product_id"],
                 cl.get("taxonomy_node_id")))
        except Exception as e:
            print(f"  Error inserting classification {cl.get('id')}: {e}")
    # Record initial seed as a batch
    now = datetime.now().isoformat()
    c.execute("INSERT INTO core_batch (name, status, total_products, processed, succeeded, failed, started_at, finished_at) VALUES (?,?,?,?,?,?,?,?)",
              ("Initial Seed (JSON chunks)", "completed", len(prods), len(prods), len(prods), 0, now, now))
    c.commit()
    c.close()
    print(f"Seeded {len(prods)} products, {len(nodes or [])} nodes, {len(cls)} classifications")

init_db()

# ---------------------------------------------------------------------------
# CSS & HTML helpers
# ---------------------------------------------------------------------------
CSS = """
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,sans-serif;background:#f5f5f5;color:#333}
.nb{background:#1a1a2e;color:white;padding:.8rem 2rem;display:flex;align-items:center;gap:1.5rem;flex-wrap:wrap}
.nb a{color:#a0a0c0;text-decoration:none;font-size:.85rem;padding:4px 8px;border-radius:4px}
.nb a:hover,.nb a.on{color:white;background:rgba(255,255,255,.1)}
.nb b{font-size:1rem;margin-right:1rem}
.wr{max-width:1400px;margin:0 auto;padding:1.5rem}
.cds{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1rem;margin-bottom:1.5rem}
.cd{background:white;border-radius:8px;padding:1rem;box-shadow:0 1px 3px rgba(0,0,0,.08)}
.cd h3{font-size:.75rem;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:.3rem}
.cd .v{font-size:1.6rem;font-weight:700}
table{width:100%;border-collapse:collapse;background:white;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08)}
th,td{padding:.6rem .8rem;text-align:left;border-bottom:1px solid #eee;font-size:.82rem}
th{background:#f8f9fa;font-weight:600;color:#555}
tr:hover{background:#f8f9fa}
.btn{display:inline-block;padding:5px 12px;border-radius:6px;border:none;cursor:pointer;font-size:.78rem;font-weight:500;text-decoration:none}
.bp{background:#007bff;color:white}.bsg{background:#28a745;color:white}
.bda{background:#dc3545;color:white}.bse{background:#6c757d;color:white}
.bwy{background:#ffc107;color:#333}.binf{background:#17a2b8;color:white}
.sb{display:flex;gap:.5rem;margin-bottom:1rem;flex-wrap:wrap;align-items:center}
.sb input,.sb select{padding:5px 10px;border:1px solid #ddd;border-radius:6px;font-size:.82rem}
.sb input{flex:1;min-width:180px}
.cw{font-weight:600;color:#28a745}.cm{font-weight:600;color:#ffc107}.cl{font-weight:600;color:#dc3545}
.pi{width:45px;height:45px;object-fit:cover;border-radius:4px}
.dg{display:grid;grid-template-columns:1fr 1fr;gap:1.2rem}
.ds{background:white;border-radius:8px;padding:1rem;box-shadow:0 1px 3px rgba(0,0,0,.08)}
.ds h3{margin-bottom:.6rem;font-size:.9rem;color:#333}
.ar{display:flex;justify-content:space-between;padding:.25rem 0;border-bottom:1px solid #f5f5f5;font-size:.82rem}
.ai{padding:.3rem 0;border-bottom:1px solid #f5f5f5;font-size:.82rem}
.pbar{height:8px;background:#eee;border-radius:4px;overflow:hidden;margin-top:.3rem}
.pfill{height:100%;border-radius:4px;transition:width .3s}
.cbar{display:flex;align-items:center;gap:.5rem;margin:.2rem 0;font-size:.8rem}
.cbarf{height:14px;background:#007bff;border-radius:3px;min-width:2px}
@media(max-width:768px){.dg{grid-template-columns:1fr}}
"""

def pg(title, body):
    nav = "<nav class='nb'><b>Taxonomy Classifier</b>"
    routes = [("/", "Dashboard"), ("/products", "Products"), ("/batches", "Batches"), ("/import", "Import")]
    for href, label in routes:
        cls = " on" if label.lower() in title.lower() else ""
        nav += f"<a href='{href}' class='{cls.strip()}'>{label}</a>"
    nav += "</nav>"
    html = f"<!DOCTYPE html><html><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width,initial-scale=1'><meta http-equiv='Cache-Control' content='no-cache, no-store, must-revalidate'><meta http-equiv='Pragma' content='no-cache'><meta http-equiv='Expires' content='0'><title>{title}</title><style>{CSS}</style></head><body>{nav}<div class='wr'>{body}</div></body></html>"
    resp = app.make_response(html)
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

def badge(status):
    m = {"approved": "bsg", "needs_review": "bwy", "rejected": "bda", "auto_classified": "binf", "pending": "bse"}
    return f"<span class='{m.get(status, 'bse')}'>{status}</span>"

def confcls(c):
    if c >= 0.7: return "cw"
    if c >= 0.4: return "cm"
    return "cl"

def safe_json(s, default=None):
    if not s: return default
    try: return json.loads(s)
    except: return default

# ---------------------------------------------------------------------------
# Taxonomy data (in-memory for classifier)
# ---------------------------------------------------------------------------
def _load_taxonomy():
    db = get_db()
    rows = db.execute("SELECT * FROM core_taxonomynode").fetchall()
    db.close()
    nodes = {}
    for r in rows:
        nodes[r["id"]] = dict(r)
        if nodes[r["id"]].get("attributes"):
            nodes[r["id"]]["attributes"] = safe_json(nodes[r["id"]]["attributes"], [])
        else:
            nodes[r["id"]]["attributes"] = []
    return nodes

TAXONOMY = None
def get_taxonomy():
    global TAXONOMY
    if TAXONOMY is None:
        TAXONOMY = _load_taxonomy()
    return TAXONOMY

def reload_taxonomy():
    global TAXONOMY
    TAXONOMY = _load_taxonomy()

# ---------------------------------------------------------------------------
# Classifier
# ---------------------------------------------------------------------------
COLORS = ["white","black","brown","gray","grey","beige","navy","blue","red","green","gold","silver","natural","walnut","oak","espresso","tan","cream","pink","purple","yellow","orange","charcoal","mahogany","teak","acacia"]
MATERIAL_KW = {"wood":"Wood","metal":"Metal","steel":"Steel","iron":"Iron","aluminum":"Aluminum","fabric":"Fabric","leather":"Leather","velvet":"Velvet","linen":"Linen","cotton":"Cotton","polyester":"Polyester","plastic":"Plastic","marble":"Marble","glass":"Glass","rattan":"Rattan","wicker":"Wicker","bamboo":"Bamboo","mango wood":"Mango Wood","rubberwood":"Rubberwood","acacia":"Acacia","eucalyptus":"Eucalyptus","teak":"Teak","engineered wood":"Engineered Wood","mdf":"MDF","particleboard":"Particleboard","ceramic":"Ceramic","concrete":"Concrete","stone":"Stone","resin":"Resin","boucle":"Boucle","chenille":"Chenille","microfiber":"Microfiber","faux leather":"Faux Leather","bonded leather":"Bonded Leather","top grain leather":"Top Grain Leather","solid wood":"Solid Wood"}
SHAPES = {"round":"Round","square":"Square","rectangular":"Rectangular","oval":"Oval","c-shaped":"C-Shaped","half-moon":"Half Moon"}

def detect_color(text):
    t = text.lower()
    for c in COLORS:
        if c in t: return c.title()
    return ""

def detect_materials(text):
    t = text.lower()
    found = []
    for kw, val in MATERIAL_KW.items():
        if kw in t and val not in found: found.append(val)
    return found[:3]

def detect_shape(text):
    t = text.lower()
    for kw, val in SHAPES.items():
        if kw in t: return val
    return ""

def detect_style(text):
    t = text.lower()
    for kw in ["modern","contemporary","mid-century","farmhouse","industrial","scandinavian","bohemian","traditional","rustic","minimalist","transitional","coastal","art deco","vintage","retro","glam","victorian","colonial","japanese","asian","moroccan","mediterranean","loft","nordic","zen","baroque"]:
        if kw in t: return kw.title()
    return ""

def extract_brand(product_name, collection_name):
    if product_name and " by " in product_name:
        parts = product_name.rsplit(" by ", 1)
        if len(parts) == 2 and len(parts[1]) < 40:
            return parts[1].strip()
    if collection_name and collection_name.lower() not in ("", "n/a", "none"):
        return collection_name.strip()
    return ""

def classify_product(product_name, source_category, source_subcategory, description="", materials="", brand=""):
    tax = get_taxonomy()
    name_lower = (product_name or "").lower()
    desc_lower = (description or "").lower()
    mat_lower = (materials or "").lower()
    combined = f"{name_lower} {desc_lower} {mat_lower}"

    scores = []
    # Category map
    cat_map = {
        ("Living Room","Sofas and Armchairs"):2,("Living Room","Coffee Tables"):4,("Living Room","TV Stands"):5,
        ("Living Room","Bookshelves"):6,("Living Room","Ottomans"):7,("Bedroom","Beds"):9,("Bedroom","Dressers"):10,
        ("Bedroom","Nightstands"):11,("Bedroom","Wardrobes"):12,("Office","Desks"):14,("Office","Chairs"):15,
        ("Office","Filing Cabinets"):16,("Dining","Tables"):18,("Dining","Chairs"):19,("Dining","Bar Stools"):20,
        ("Dining","Buffets and Sideboards"):21,("Dining","Dining Sets"):22,("Outdoor","Patio Furniture"):24,
        ("Outdoor","Outdoor Dining"):25,("Outdoor","Outdoor Seating"):26,("Bath","Bathroom Vanities"):28,
        ("Bath","Bathroom Storage"):29,("Bath","Bathroom Mirrors"):30,
    }
    key = (source_category, source_subcategory)
    if key in cat_map:
        scores.append((cat_map[key], 0.95))

    # Keyword scoring
    for nid, node in tax.items():
        if node["level"] == 0:
            continue
        kws = [k.strip() for k in (node.get("keywords") or "").split(",") if k.strip()]
        matched = sum(1 for k in kws if k in combined)
        if matched > 0:
            sc = min(0.6 + 0.08 * matched, 0.94)
            scores.append((nid, sc))

    if not scores:
        return 1, 0.30, [], {}

    scores.sort(key=lambda x: -x[1])
    best_nid, best_score = scores[0]

    # Parent boost
    parent_id = tax.get(best_nid, {}).get("parent_id")
    if parent_id:
        parent = tax.get(parent_id, {})
        pkws = [k.strip() for k in (parent.get("keywords") or "").split(",") if k.strip()]
        if sum(1 for k in pkws if k in combined) > 0:
            best_score = min(best_score + 0.05, 0.99)

    # Alternatives
    alternatives = []
    seen = {best_nid}
    for nid, sc in scores[1:]:
        if nid not in seen and len(alternatives) < 3:
            n = tax.get(nid, {})
            alternatives.append({"name": n.get("name",""), "confidence": round(sc,3), "path": _build_path(nid, tax)})
            seen.add(nid)

    # Detect attributes
    attrs = {}
    color = detect_color(combined)
    if color: attrs["Color"] = color
    mats = detect_materials(combined)
    if mats: attrs["Material"] = ", ".join(mats)
    shape = detect_shape(combined)
    if shape: attrs["Shape"] = shape
    style = detect_style(combined)
    if style: attrs["Style"] = style

    return best_nid, round(best_score, 3), alternatives, attrs

def _build_path(node_id, tax):
    parts = []
    nid = node_id
    while nid:
        n = tax.get(nid)
        if not n: break
        parts.append(n["name"])
        nid = n.get("parent_id")
    return " > ".join(reversed(parts))

# ---------------------------------------------------------------------------
# Routes: Dashboard
# ---------------------------------------------------------------------------
@app.route("/")
def dashboard():
    db = get_db()
    total = db.execute("SELECT COUNT(*) FROM core_product").fetchone()[0]
    classified = db.execute("SELECT COUNT(*) FROM core_classification").fetchone()[0]
    nr = db.execute("SELECT COUNT(*) FROM core_classification WHERE needs_review=1").fetchone()[0]
    ap = db.execute("SELECT COUNT(*) FROM core_classification WHERE status='approved'").fetchone()[0]
    ac = db.execute("SELECT AVG(confidence) FROM core_classification").fetchone()[0] or 0
    sc = dict(db.execute("SELECT status, COUNT(*) FROM core_classification GROUP BY status").fetchall())
    tc = db.execute("""SELECT t.name, COUNT(*) as c FROM core_classification cl
        JOIN core_taxonomynode t ON cl.taxonomy_node_id=t.id
        GROUP BY t.name ORDER BY c DESC LIMIT 10""").fetchall()
    with_img = db.execute("SELECT COUNT(*) FROM core_product WHERE image_url IS NOT NULL AND image_url != ''").fetchone()[0]
    without_img = total - with_img
    db.close()
    s = "<h2 style='margin-bottom:1rem'>Dashboard</h2><div class='cds'>"
    s += f"<div class='cd'><h3>Total Products</h3><div class='v'>{total}</div></div>"
    s += f"<div class='cd'><h3>Classified</h3><div class='v'>{classified}</div></div>"
    s += f"<div class='cd'><h3>Needs Review</h3><div class='v' style='color:#ffc107'>{nr}</div></div>"
    s += f"<div class='cd'><h3>Approved</h3><div class='v' style='color:#28a745'>{ap}</div></div>"
    s += f"<div class='cd'><h3>Avg Confidence</h3><div class='v'>{ac:.1%}</div></div>"
    s += f"<div class='cd'><h3>With Image</h3><div class='v'>{with_img}</div></div>"
    s += f"<div class='cd'><h3>No Image</h3><div class='v'>{without_img}</div></div></div>"
    s += "<div class='dg'><div class='ds'><h3>Status Breakdown</h3>"
    for st, c in sc.items():
        w = max(c * 150 // max(total, 1), 2)
        s += f"<div class='cbar'><span style='width:130px'>{st}</span><div class='cbarf' style='width:{w}px'></div><span>{c}</span></div>"
    s += "</div><div class='ds'><h3>Top Categories</h3>"
    mx = tc[0][1] if tc else 1
    for nm, c in tc:
        w = max(c * 150 // mx, 2)
        s += f"<div class='cbar'><span style='width:180px'>{nm}</span><div class='cbarf' style='width:{w}px'></div><span>{c}</span></div>"
    s += "</div></div>"
    s += "<div style='margin-top:1.2rem'><a href='/products?status=needs_review' class='btn bwy' style='text-decoration:none'>Review Flagged</a> "
    s += "<a href='/products' class='btn bse' style='text-decoration:none;margin-left:.5rem'>View All</a> "
    s += "<a href='/import' class='btn bp' style='text-decoration:none;margin-left:.5rem'>Import Products</a></div>"
    return pg("Dashboard", s)

# ---------------------------------------------------------------------------
# Routes: Products
# ---------------------------------------------------------------------------
@app.route("/products")
def product_list():
    db = get_db()
    st = request.args.get("status", "")
    q = request.args.get("q", "")
    page = max(int(request.args.get("page", 1)), 1)
    per_page = 50
    offset = (page - 1) * per_page

    where = "WHERE 1=1"
    params = []
    if st:
        where += " AND cl.status=?"
        params.append(st)
    if q:
        where += " AND (p.product_name LIKE ? OR p.product_number LIKE ? OR p.source_category LIKE ? OR p.brand LIKE ?)"
        params.extend([f"%{q}%"] * 4)

    total = db.execute(f"SELECT COUNT(*) FROM core_classification cl JOIN core_product p ON cl.product_id=p.id {where}", params).fetchone()[0]
    rows = db.execute(f"""SELECT cl.*, p.product_number, p.product_name, p.image_url, p.source_category, p.brand, t.name as tax_name
        FROM core_classification cl JOIN core_product p ON cl.product_id=p.id
        LEFT JOIN core_taxonomynode t ON cl.taxonomy_node_id=t.id {where}
        ORDER BY cl.confidence DESC LIMIT ? OFFSET ?""", params + [per_page, offset]).fetchall()
    db.close()

    total_pages = max((total + per_page - 1) // per_page, 1)
    opts = ""
    for val, lbl in [("pending","Pending"),("auto_classified","Auto Classified"),("needs_review","Needs Review"),("approved","Approved"),("rejected","Rejected")]:
        sel = "selected" if val == st else ""
        opts += f"<option value='{val}' {sel}>{lbl}</option>"

    s = "<h2 style='margin-bottom:1rem'>Product Classifications</h2>"
    s += f"<form class='sb' method='get'><input type='text' name='q' placeholder='Search name, SKU, brand...' value='{q}'><select name='status'><option value=''>All Status</option>{opts}</select><button type='submit' class='btn bp'>Filter</button></form>"
    s += f"<p style='font-size:.8rem;color:#888;margin-bottom:.5rem'>Showing {offset+1}-{min(offset+per_page,total)} of {total} products</p>"
    s += "<table><thead><tr><th>Image</th><th>SKU</th><th>Name</th><th>Brand</th><th>Category</th><th>Confidence</th><th>Status</th><th></th></tr></thead><tbody>"
    for r in rows:
        img_url = r["image_url"] or ""
        img = f"<img src='{img_url}' class='pi' loading='lazy' onerror=\"this.style.display='none'\">" if img_url else ""
        nm = (r["product_name"] or "-")[:50]
        brand = (r["brand"] or "-")[:20]
        s += f"<tr><td>{img}</td><td><code>{r['product_number']}</code></td><td>{nm}</td><td>{brand}</td><td>{r['tax_name'] or '-'}</td><td><span class='{confcls(r['confidence'])}'>{r['confidence']*100:.1f}%</span></td><td>{badge(r['status'])}</td><td><a href='/products/{r['id']}' class='btn bp'>View</a></td></tr>"
    s += "</tbody></table>"
    # Pagination
    s += "<div style='margin-top:1rem;display:flex;gap:.5rem'>"
    if page > 1:
        s += f"<a href='?page={page-1}&status={st}&q={q}' class='btn bse'>Prev</a>"
    s += f"<span style='padding:5px 10px;font-size:.82rem'>Page {page}/{total_pages}</span>"
    if page < total_pages:
        s += f"<a href='?page={page+1}&status={st}&q={q}' class='btn bse'>Next</a>"
    s += "</div>"
    return pg("Products", s)

@app.route("/products/<int:cid>")
def product_detail(cid):
    db = get_db()
    r = db.execute("""SELECT cl.*, p.product_number, p.product_name, p.product_description, p.image_url,
        p.source_category, p.source_subcategory, p.collection_name, p.materials, p.product_dimensions,
        p.product_weight, p.country_of_origin, p.model_number, p.brand, p.product_type, p.all_image_urls,
        t.name as tax_name, t.attributes as node_attributes
        FROM core_classification cl JOIN core_product p ON cl.product_id=p.id
        LEFT JOIN core_taxonomynode t ON cl.taxonomy_node_id=t.id WHERE cl.id=?""", [cid]).fetchone()
    if not r:
        db.close()
        return "Not found", 404
    alts = safe_json(r["alternatives"], [])
    attrs = safe_json(r["detected_attributes"], {})
    node_attrs = safe_json(r["node_attributes"], [])
    all_imgs = safe_json(r["all_image_urls"], [])
    db.close()

    c = r["confidence"]
    cp = f"{c*100:.1f}%"
    bg = "#28a745" if c >= 0.7 else ("#ffc107" if c >= 0.4 else "#dc3545")
    img_url = r["image_url"] or ""
    img = f"<img src='{img_url}' style='max-width:200px;max-height:200px;border-radius:8px' onerror=\"this.style.display='none'\">" if img_url else ""
    desc = (r["product_description"] or "")[:500]
    mat = f"<p style='font-size:.8rem;color:#666;margin-top:.3rem'>Materials: {r['materials']}</p>" if r["materials"] else ""
    dim = f"<p style='font-size:.8rem;color:#666'>Dimensions: {(r['product_dimensions'] or '')[:100]}</p>" if r["product_dimensions"] else ""

    # Attributes panel
    ah = ""
    for k, v in attrs.items():
        ah += f"<div class='ar'><span>{k}</span><strong>{v}</strong></div>"
    if not ah:
        ah = "<p style='color:#888;font-size:.85rem'>None detected.</p>"

    # Expected attributes from taxonomy
    expected = ""
    if node_attrs:
        expected = "<div class='ds' style='margin-top:1rem'><h3>Expected Category Attributes</h3>"
        for a in node_attrs:
            val = attrs.get(a, "<em style='color:#ccc'>not detected</em>")
            expected += f"<div class='ar'><span>{a}</span><strong>{val}</strong></div>"
        expected += "</div>"

    # Alternatives
    al = ""
    for a in alts:
        al += f"<div class='ai'><strong>{a['name']}</strong> <span style='color:#888;font-size:.78rem'>{a.get('path','')}</span> <span class='{confcls(a['confidence'])}'>{a['confidence']*100:.1f}%</span></div>"
    als = f"<div class='ds' style='margin-top:1rem'><h3>Alternative Categories</h3>{al}</div>" if al else ""

    # Gallery
    gallery = ""
    if all_imgs and len(all_imgs) > 1:
        gallery = "<div style='display:flex;gap:.5rem;flex-wrap:wrap;margin-top:.5rem'>"
        for url in all_imgs[:6]:
            gallery += f"<img src='{url}' style='width:60px;height:60px;object-fit:cover;border-radius:4px' onerror=\"this.style.display='none'\">"
        gallery += "</div>"

    wt = f"{r['product_weight']} lbs" if r["product_weight"] else "-"
    src = f"{r['source_category'] or '-'} > {r['source_subcategory'] or '-'}"
    brand = r["brand"] or "-"
    ptype = r["product_type"] or "-"

    s = "<div style='margin-bottom:1rem'><a href='/products' style='color:#007bff;text-decoration:none;font-size:.85rem'>&larr; Back</a></div>"
    s += "<div style='display:flex;gap:1.5rem;align-items:flex-start;flex-wrap:wrap'>"
    s += f"<div style='flex:2;min-width:300px'><div class='ds'><div style='display:flex;gap:1rem;align-items:flex-start'>{img}<div><h3 style='margin-bottom:.3rem'>{r['product_name'] or ''}</h3><p style='color:#888;font-size:.82rem'>{r['product_number']} | {r['model_number'] or ''}</p><p style='font-size:.82rem;margin-top:.3rem'><b>Brand:</b> {brand} | <b>Type:</b> {ptype}</p><p style='font-size:.82rem;margin-top:.4rem'>{desc}</p>{mat}{dim}</div></div>{gallery}</div></div>"
    s += "<div style='flex:1;min-width:280px'>"
    s += f"<div class='ds'><h3>Classification</h3><div class='ar'><span>Category</span><strong>{r['tax_name'] or 'Unclassified'}</strong></div><div class='ar'><span>Confidence</span><span class='{confcls(c)}'>{cp}</span></div><div class='ar'><span>Status</span>{badge(r['status'])}</div>"
    if r["review_reason"]:
        s += f"<div class='ar'><span>Reason</span><span>{r['review_reason']}</span></div>"
    s += f"<div class='pbar'><div class='pfill' style='width:{c*100:.0f}%;background:{bg}'></div></div></div>"
    s += f"<div class='ds' style='margin-top:1rem'><h3>Detected Attributes</h3>{ah}</div>"
    s += expected
    s += als
    s += f"<div class='ds' style='margin-top:1rem'><h3>Actions</h3>"
    s += f"<form method='post' action='/products/{cid}/approve' style='display:inline'><button class='btn bsg'>Approve</button></form> "
    s += f"<form method='post' action='/products/{cid}/reject' style='display:inline'><button class='btn bda'>Reject</button></form> "
    s += f"<a href='/products/{cid}/reclassify' class='btn bwy' style='text-decoration:none'>Reclassify</a></div>"
    s += f"<div class='ds' style='margin-top:1rem'><h3>Product Info</h3>"
    s += f"<div class='ar'><span>Source</span><span>{src}</span></div>"
    s += f"<div class='ar'><span>Collection</span><span>{r['collection_name'] or '-'}</span></div>"
    s += f"<div class='ar'><span>Country</span><span>{r['country_of_origin'] or '-'}</span></div>"
    s += f"<div class='ar'><span>Weight</span><span>{wt}</span></div></div></div></div>"
    return pg(f"{r['product_number']} - Detail", s)

@app.route("/products/<int:cid>/approve", methods=["POST"])
def approve(cid):
    db = get_db()
    db.execute("UPDATE core_classification SET status='approved', needs_review=0, reviewed_at=? WHERE id=?",
               (datetime.now().isoformat(), cid))
    db.commit()
    db.close()
    return redirect(f"/products/{cid}")

@app.route("/products/<int:cid>/reject", methods=["POST"])
def reject(cid):
    db = get_db()
    db.execute("UPDATE core_classification SET status='rejected', reviewed_at=? WHERE id=?",
               (datetime.now().isoformat(), cid))
    db.commit()
    db.close()
    return redirect(f"/products/{cid}")

@app.route("/products/<int:cid>/reclassify")
def reclassify(cid):
    db = get_db()
    p = db.execute("SELECT * FROM core_product WHERE id=?", [cid]).fetchone()
    if not p:
        db.close()
        return "Product not found", 404
    text = f"{p['product_name']} {p['product_description']} {p['bullets']} {p['source_category']} {p['source_subcategory']} {p['collection_name']} {p['materials']} {p['brand'] or ''} {p['product_type'] or ''}"
    node_id, confidence, alternatives, attrs = classify_product(
        p["product_name"], p["source_category"], p["source_subcategory"],
        p["product_description"], p["materials"], p["brand"] or "")
    db.execute("""UPDATE core_classification SET taxonomy_node_id=?, confidence=?, alternatives=?,
        detected_attributes=?, status='auto_classified', needs_review=?, review_reason=? WHERE product_id=?""",
        (node_id, confidence, json.dumps(alternatives), json.dumps(attrs),
         confidence < 0.4, "" if confidence >= 0.4 else "Low confidence", cid))
    db.commit()
    db.close()
    return redirect(f"/products/{cid}")

# ---------------------------------------------------------------------------
# Routes: Batch operations
# ---------------------------------------------------------------------------
@app.route("/batch/approve", methods=["POST"])
def batch_approve():
    ids = request.form.getlist("ids")
    if not ids:
        ids_str = request.form.get("ids_csv", "")
        ids = [x.strip() for x in ids_str.split(",") if x.strip()]
    if not ids:
        return jsonify({"error": "No IDs provided"}), 400
    db = get_db()
    placeholders = ",".join("?" * len(ids))
    now = datetime.now().isoformat()
    db.execute(f"UPDATE core_classification SET status='approved', needs_review=0, reviewed_at=? WHERE id IN ({placeholders})",
               [now] + [int(i) for i in ids])
    db.commit()
    updated = db.execute(f"SELECT COUNT(*) FROM core_classification WHERE id IN ({placeholders})",
                         [int(i) for i in ids]).fetchone()[0]
    db.close()
    if request.headers.get("Accept") == "application/json":
        return jsonify({"approved": updated})
    return redirect("/products")

@app.route("/batch/reject", methods=["POST"])
def batch_reject():
    ids = request.form.getlist("ids")
    if not ids:
        ids_str = request.form.get("ids_csv", "")
        ids = [x.strip() for x in ids_str.split(",") if x.strip()]
    if not ids:
        return jsonify({"error": "No IDs provided"}), 400
    db = get_db()
    placeholders = ",".join("?" * len(ids))
    now = datetime.now().isoformat()
    db.execute(f"UPDATE core_classification SET status='rejected', reviewed_at=? WHERE id IN ({placeholders})",
               [now] + [int(i) for i in ids])
    db.commit()
    updated = db.execute(f"SELECT COUNT(*) FROM core_classification WHERE id IN ({placeholders})",
                         [int(i) for i in ids]).fetchone()[0]
    db.close()
    if request.headers.get("Accept") == "application/json":
        return jsonify({"rejected": updated})
    return redirect("/products")

@app.route("/batch/reclassify", methods=["POST"])
def batch_reclassify():
    ids = request.form.getlist("ids")
    if not ids:
        ids_str = request.form.get("ids_csv", "")
        ids = [x.strip() for x in ids_str.split(",") if x.strip()]
    if not ids:
        return jsonify({"error": "No IDs provided"}), 400
    db = get_db()
    reloaded = False
    count = 0
    for cid in ids:
        try:
            p = db.execute("SELECT * FROM core_product WHERE id=?", [int(cid)]).fetchone()
            if not p:
                continue
            if not reloaded:
                reload_taxonomy()
                reloaded = True
            node_id, confidence, alternatives, attrs = classify_product(
                p["product_name"], p["source_category"], p["source_subcategory"],
                p["product_description"], p["materials"], p["brand"] or "")
            db.execute("""UPDATE core_classification SET taxonomy_node_id=?, confidence=?, alternatives=?,
                detected_attributes=?, status=?, needs_review=?, review_reason=? WHERE product_id=?""",
                (node_id, confidence, json.dumps(alternatives), json.dumps(attrs),
                 "auto_classified" if confidence >= 0.4 else "needs_review",
                 confidence < 0.4, "" if confidence >= 0.4 else "Low confidence", int(cid)))
            count += 1
        except Exception as e:
            print(f"  Reclassify error for {cid}: {e}")
    db.commit()
    db.close()
    if request.headers.get("Accept") == "application/json":
        return jsonify({"reclassified": count})
    return redirect("/products")

# ---------------------------------------------------------------------------
# Routes: Import
# ---------------------------------------------------------------------------
def _read_xlsx_file(file_obj):
    """Read products from an uploaded .xlsx file."""
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl is required for .xlsx support: pip install openpyxl"
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_obj.read())
        tmp_path = tmp.name
    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Build header index map (case-insensitive)
    hmap = {}
    for i, h in enumerate(headers):
        if h:
            hmap[str(h).strip().lower()] = i
    products = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if row[0] is None:
                continue
            # Collect image URLs from columns 21-40
            images = []
            for ci in range(20, min(39, len(row))):
                if row[ci]:
                    images.append(str(row[ci]))
            product = {
                "product_number": str(row[hmap.get("product_number", hmap.get("sku", 0))] or "") if "product_number" in hmap or "sku" in hmap else str(row[0] or ""),
                "product_name": str(row[hmap.get("product_name", hmap.get("title", 7))] or "") if "product_name" in hmap or "title" in hmap else str(row[7] or ""),
                "product_description": str(row[hmap.get("product_description", hmap.get("description", 8))] or "") if "product_description" in hmap or "description" in hmap else str(row[8] or ""),
                "image_url": str(row[hmap.get("image_url", hmap.get("image", 21))] or "") if "image_url" in hmap or "image" in hmap else (str(row[21]) if len(row) > 21 and row[21] else ""),
                "all_image_urls": json.dumps(images),
                "source_category": str(row[hmap.get("source_category", hmap.get("category", 2))] or "") if "source_category" in hmap or "category" in hmap else str(row[2] or ""),
                "source_subcategory": str(row[hmap.get("source_subcategory", hmap.get("subcategory", 3))] or "") if "source_subcategory" in hmap or "subcategory" in hmap else str(row[3] or ""),
                "collection_name": str(row[hmap.get("collection_name", hmap.get("collection", 4))] or "") if "collection_name" in hmap or "collection" in hmap else str(row[4] or ""),
                "materials": str(row[hmap.get("materials", hmap.get("material", 12))] or "") if "materials" in hmap or "material" in hmap else (str(row[12]) if len(row) > 12 and row[12] else ""),
                "product_dimensions": str(row[hmap.get("product_dimensions", hmap.get("dimensions", 13))] or "") if "product_dimensions" in hmap or "dimensions" in hmap else (str(row[13]) if len(row) > 13 and row[13] else ""),
                "product_weight": None,
                "country_of_origin": str(row[hmap.get("country_of_origin", hmap.get("country", 17))] or "") if "country_of_origin" in hmap or "country" in hmap else (str(row[17]) if len(row) > 17 and row[17] else ""),
                "product_url": str(row[hmap.get("product_url", hmap.get("url", 47))] or "") if "product_url" in hmap or "url" in hmap else (str(row[47]) if len(row) > 47 and row[47] else ""),
                "model_number": str(row[hmap.get("model_number", 1)] or "") if "model_number" in hmap else (str(row[1]) if len(row) > 1 and row[1] else ""),
                "bullets": str(row[hmap.get("bullets", 9)] or "") if "bullets" in hmap else (str(row[9]) if len(row) > 9 and row[9] else ""),
                "brand": str(row[hmap.get("brand", "")] or "") if "brand" in hmap else "",
                "product_type": str(row[hmap.get("product_type", "")] or "") if "product_type" in hmap else "",
            }
            # Parse weight
            wcol = hmap.get("product_weight", hmap.get("weight", 11))
            if wcol is not None and wcol < len(row) and row[wcol]:
                try:
                    product["product_weight"] = float(row[wcol])
                except (ValueError, TypeError):
                    pass
            # Derive brand/product_type if missing
            if not product["brand"] and product["product_name"] and " by " in product["product_name"]:
                product["brand"] = product["product_name"].rsplit(" by ", 1)[1].strip()
            if not product["product_type"]:
                product["product_type"] = product["source_subcategory"]
            products.append(product)
        except Exception:
            continue
    wb.close()
    os.unlink(tmp_path)
    return products, None


def _read_csv_file(file_obj):
    """Read products from an uploaded .csv file."""
    content = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    products = []
    for row in reader:
        try:
            product = {
                "product_number": row.get("product_number", row.get("SKU", "")),
                "product_name": row.get("product_name", row.get("title", "")),
                "product_description": row.get("product_description", row.get("description", "")),
                "image_url": row.get("image_url", row.get("image", "")),
                "all_image_urls": "[]",
                "source_category": row.get("source_category", row.get("category", "")),
                "source_subcategory": row.get("source_subcategory", row.get("subcategory", "")),
                "collection_name": row.get("collection_name", row.get("collection", "")),
                "materials": row.get("materials", row.get("material", "")),
                "product_dimensions": row.get("product_dimensions", row.get("dimensions", "")),
                "product_weight": float(row.get("product_weight", row.get("weight", 0)) or 0) or None,
                "country_of_origin": row.get("country_of_origin", row.get("country", "")),
                "product_url": row.get("product_url", row.get("url", "")),
                "model_number": row.get("model_number", ""),
                "bullets": row.get("bullets", ""),
                "brand": row.get("brand", ""),
                "product_type": row.get("product_type", ""),
            }
            if not product["brand"]:
                if product["product_name"] and " by " in product["product_name"]:
                    product["brand"] = product["product_name"].rsplit(" by ", 1)[1].strip()
            if not product["product_type"]:
                product["product_type"] = product["source_subcategory"]
            products.append(product)
        except Exception:
            continue
    return products, None


@app.route("/import", methods=["GET", "POST"])
def import_page():
    if request.method == "GET":
        s = "<h2 style='margin-bottom:1rem'>Import Products</h2>"
        s += "<div class='ds'><h3>Upload CSV or Excel File</h3>"
        s += "<form method='post' enctype='multipart/form-data' style='margin-top:.5rem'>"
        s += "<input type='file' name='file' accept='.csv,.xlsx,.xls' style='margin-bottom:.5rem'><br>"
        s += "<button type='submit' class='btn bp'>Import & Classify</button></form>"
        s += "<p style='font-size:.8rem;color:#888;margin-top:.5rem'>Supported formats: .csv, .xlsx, .xls</p>"
        s += "<p style='font-size:.8rem;color:#888'>Columns: product_number, product_name, source_category, source_subcategory, materials, brand, product_type, image_url, product_description</p>"
        s += "</div>"
        return pg("Import", s)

    file = request.files.get("file")
    if not file or not file.filename:
        return "No file uploaded", 400

    filename = file.filename.lower()
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            products, err = _read_xlsx_file(file)
            if err:
                return f"XLSX error: {err}", 500
        elif filename.endswith(".csv"):
            products, err = _read_csv_file(file)
            if err:
                return f"CSV error: {err}", 500
        else:
            return "Unsupported file format. Use .csv, .xlsx, or .xls", 400

        if not products:
            return "No valid products found in file", 400

        db = get_db()

        # Clear old data on new import
        db.execute("DELETE FROM core_classification")
        db.execute("DELETE FROM core_product")
        db.execute("DELETE FROM core_batch")
        db.commit()

        reload_taxonomy()
        inserted = 0
        errors_list = []

        for i, p in enumerate(products):
            try:
                pid = i + 1
                p["id"] = pid
                db.execute("INSERT OR IGNORE INTO core_product VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (pid, p["product_number"], p["product_name"], p["product_description"],
                     p["image_url"], p["all_image_urls"], p["source_category"], p["source_subcategory"],
                     p["collection_name"], p["materials"], p["product_dimensions"], p["product_weight"],
                     p["country_of_origin"], p["product_url"], p["model_number"], p["bullets"],
                     p["brand"], p["product_type"]))
                text = f"{p['product_name']} {p['product_description']} {p['bullets']} {p['source_category']} {p['source_subcategory']} {p['collection_name']} {p['materials']} {p['brand']} {p['product_type']}"
                node_id, confidence, alternatives, attrs = classify_product(
                    p["product_name"], p["source_category"], p["source_subcategory"],
                    p["product_description"], p["materials"], p["brand"])
                has_img = bool(p.get("image_url"))
                has_desc = bool(p.get("product_description", "").strip())
                if confidence < 0.4:
                    status, reason, nr = "needs_review", "Low confidence", True
                elif not has_desc and not has_img:
                    status, reason, nr = "needs_review", "Missing description and image", True
                elif not has_desc:
                    status, reason, nr = "needs_review", "Missing description", True
                else:
                    status, reason, nr = "auto_classified", "", False
                db.execute("INSERT INTO core_classification VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (pid, confidence, status, json.dumps(alternatives), json.dumps(attrs),
                     nr, reason, datetime.now().isoformat(), None, "", pid, node_id))
                inserted += 1
            except Exception as e:
                errors_list.append(f"Row {i+1}: {str(e)}")

        db.commit()

        # Record batch
        try:
            now_iso = datetime.now().isoformat()
            db.execute("INSERT INTO core_batch (name, status, total_products, processed, succeeded, failed, started_at, finished_at) VALUES (?,?,?,?,?,?,?,?)",
                      (f"Import: {file.filename}", "completed", len(products), inserted, inserted, len(errors_list), now_iso, now_iso))
            db.commit()
        except Exception:
            pass

        db.close()

        s = "<h2 style='margin-bottom:1rem'>Import Complete</h2><div class='ds'>"
        s += f"<p>Imported <strong>{inserted}</strong> products</p>"
        if errors_list:
            s += f"<p style='color:#dc3545'>{len(errors_list)} errors:</p><ul>"
            for e in errors_list[:10]:
                s += f"<li style='font-size:.82rem'>{e}</li>"
            s += "</ul>"
        s += "<a href='/' class='btn bsg' style='text-decoration:none;margin-top:.5rem'>View Dashboard</a> "
        s += "<a href='/products' class='btn bp' style='text-decoration:none;margin-top:.5rem'>View Products</a> "
        s += "<a href='/import' class='btn bse' style='text-decoration:none;margin-top:.5rem'>Import More</a></div>"
        return pg("Import Results", s)
    except Exception as e:
        return f"Import error: {str(e)}", 500

# ---------------------------------------------------------------------------
# Routes: Batches
# ---------------------------------------------------------------------------
@app.route("/batches")
def batch_list():
    db = get_db()
    try:
        batches = db.execute("SELECT * FROM core_batch ORDER BY created_at DESC LIMIT 20").fetchall()
    except:
        batches = []
    db.close()
    s = "<h2 style='margin-bottom:1rem'>Batch History</h2>"
    if not batches:
        s += "<div class='ds'><p style='color:#888'>No batch processing history yet. Use the API to trigger batch operations.</p></div>"
    else:
        s += "<table><thead><tr><th>ID</th><th>Name</th><th>Status</th><th>Total</th><th>Processed</th><th>Failed</th><th>Started</th><th>Finished</th></tr></thead><tbody>"
        for b in batches:
            s += f"<tr><td>{b['id']}</td><td>{b['name']}</td><td>{badge(b['status'])}</td><td>{b['total_products']}</td><td>{b['processed']}</td><td>{b['failed']}</td><td>{b['started_at'] or '-'}</td><td>{b['finished_at'] or '-'}</td></tr>"
        s += "</tbody></table>"
    return pg("Batches", s)

# ---------------------------------------------------------------------------
# API endpoints
# ---------------------------------------------------------------------------
@app.route("/api/stats")
def api_stats():
    db = get_db()
    t = db.execute("SELECT COUNT(*) FROM core_product").fetchone()[0]
    cl = db.execute("SELECT COUNT(*) FROM core_classification").fetchone()[0]
    sc = dict(db.execute("SELECT status, COUNT(*) FROM core_classification GROUP BY status").fetchall())
    ac = db.execute("SELECT AVG(confidence) FROM core_classification").fetchone()[0] or 0
    nr = db.execute("SELECT COUNT(*) FROM core_classification WHERE needs_review=1").fetchone()[0]
    db.close()
    return jsonify({"total_products": t, "classified": cl, "status_counts": sc,
                    "avg_confidence": round(ac, 3), "needs_review": nr})

@app.route("/api/products")
def api_products():
    db = get_db()
    page = max(int(request.args.get("page", 1)), 1)
    per_page = min(int(request.args.get("per_page", 50)), 200)
    offset = (page - 1) * per_page
    total = db.execute("SELECT COUNT(*) FROM core_product").fetchone()[0]
    rows = db.execute("SELECT * FROM core_product ORDER BY id LIMIT ? OFFSET ?", [per_page, offset]).fetchall()
    db.close()
    return jsonify({"total": total, "page": page, "per_page": per_page,
                    "products": [dict(r) for r in rows]})

@app.route("/api/products/<int:pid>")
def api_product_detail(pid):
    db = get_db()
    p = db.execute("SELECT * FROM core_product WHERE id=?", [pid]).fetchone()
    if not p:
        db.close()
        return jsonify({"error": "Not found"}), 404
    cl = db.execute("SELECT * FROM core_classification WHERE product_id=?", [pid]).fetchone()
    node = None
    if cl and cl["taxonomy_node_id"]:
        node = db.execute("SELECT * FROM core_taxonomynode WHERE id=?", [cl["taxonomy_node_id"]]).fetchone()
    db.close()
    result = dict(p)
    if cl:
        result["classification"] = dict(cl)
        result["classification"]["alternatives"] = safe_json(cl["alternatives"], [])
        result["classification"]["detected_attributes"] = safe_json(cl["detected_attributes"], {})
    if node:
        result["taxonomy_node"] = dict(node)
        result["taxonomy_node"]["attributes"] = safe_json(node["attributes"], [])
    return jsonify(result)

@app.route("/api/classifications")
def api_classifications():
    db = get_db()
    status = request.args.get("status", "")
    q = request.args.get("q", "")
    where = "WHERE 1=1"
    params = []
    if status:
        where += " AND cl.status=?"
        params.append(status)
    if q:
        where += " AND (p.product_name LIKE ? OR p.product_number LIKE ?)"
        params.extend([f"%{q}%"] * 2)
    rows = db.execute(f"""SELECT cl.id, p.product_number, p.product_name, t.name as taxonomy,
        cl.confidence, cl.status, cl.needs_review, cl.detected_attributes
        FROM core_classification cl JOIN core_product p ON cl.product_id=p.id
        LEFT JOIN core_taxonomynode t ON cl.taxonomy_node_id=t.id {where}
        ORDER BY cl.confidence DESC LIMIT 100""", params).fetchall()
    db.close()
    result = []
    for r in rows:
        d = dict(r)
        d["detected_attributes"] = safe_json(d["detected_attributes"], {})
        result.append(d)
    return jsonify(result)

@app.route("/api/classifications/<int:cid>", methods=["PUT"])
def api_update_classification(cid):
    data = request.get_json(force=True, silent=True) or {}
    db = get_db()
    cl = db.execute("SELECT * FROM core_classification WHERE id=?", [cid]).fetchone()
    if not cl:
        db.close()
        return jsonify({"error": "Not found"}), 404
    updates = []
    params = []
    if "status" in data:
        updates.append("status=?")
        params.append(data["status"])
    if "taxonomy_node_id" in data:
        updates.append("taxonomy_node_id=?")
        params.append(int(data["taxonomy_node_id"]))
    if "confidence" in data:
        updates.append("confidence=?")
        params.append(float(data["confidence"]))
    if "needs_review" in data:
        updates.append("needs_review=?")
        params.append(1 if data["needs_review"] else 0)
    if "reviewer_notes" in data:
        updates.append("reviewer_notes=?")
        params.append(data["reviewer_notes"])
    if "status" in data and data["status"] in ("approved", "rejected"):
        updates.append("reviewed_at=?")
        updates.append("needs_review=0")
        params.append(datetime.now().isoformat())
    if updates:
        params.append(cid)
        db.execute(f"UPDATE core_classification SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
    updated = db.execute("SELECT * FROM core_classification WHERE id=?", [cid]).fetchone()
    db.close()
    return jsonify(dict(updated))

@app.route("/api/classifications/batch", methods=["POST"])
def api_batch_update():
    data = request.get_json(force=True, silent=True) or {}
    ids = data.get("ids", [])
    action = data.get("action", "")
    if not ids or action not in ("approve", "reject", "reclassify"):
        return jsonify({"error": "Provide ids array and action (approve/reject/reclassify)"}), 400

    db = get_db()
    now = datetime.now().isoformat()
    placeholders = ",".join("?" * len(ids))
    int_ids = [int(i) for i in ids]

    if action == "approve":
        db.execute(f"UPDATE core_classification SET status='approved', needs_review=0, reviewed_at=? WHERE id IN ({placeholders})",
                   [now] + int_ids)
    elif action == "reject":
        db.execute(f"UPDATE core_classification SET status='rejected', reviewed_at=? WHERE id IN ({placeholders})",
                   [now] + int_ids)
    elif action == "reclassify":
        reload_taxonomy()
        for cid in int_ids:
            try:
                p = db.execute("SELECT * FROM core_product WHERE id=?", [cid]).fetchone()
                if not p: continue
                node_id, confidence, alternatives, attrs = classify_product(
                    p["product_name"], p["source_category"], p["source_subcategory"],
                    p["product_description"], p["materials"], p["brand"] or "")
                db.execute("""UPDATE core_classification SET taxonomy_node_id=?, confidence=?, alternatives=?,
                    detected_attributes=?, status=?, needs_review=? WHERE product_id=?""",
                    (node_id, confidence, json.dumps(alternatives), json.dumps(attrs),
                     "auto_classified" if confidence >= 0.4 else "needs_review",
                     confidence < 0.4, cid))
            except Exception as e:
                print(f"  Batch reclassify error {cid}: {e}")
    db.commit()
    db.close()
    return jsonify({"updated": len(ids), "action": action})

@app.route("/api/taxonomy")
def api_taxonomy():
    db = get_db()
    nodes = db.execute("SELECT * FROM core_taxonomynode ORDER BY level, id").fetchall()
    db.close()
    result = []
    for n in nodes:
        d = dict(n)
        d["attributes"] = safe_json(d["attributes"], [])
        result.append(d)
    return jsonify(result)

@app.route("/api/taxonomy/<int:nid>")
def api_taxonomy_node(nid):
    db = get_db()
    node = db.execute("SELECT * FROM core_taxonomynode WHERE id=?", [nid]).fetchone()
    db.close()
    if not node:
        return jsonify({"error": "Not found"}), 404
    d = dict(node)
    d["attributes"] = safe_json(d["attributes"], [])
    return jsonify(d)

@app.route("/api/import", methods=["POST"])
def api_import():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400
    try:
        content = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        db = get_db()

        # Clear old data on new import
        db.execute("DELETE FROM core_classification")
        db.execute("DELETE FROM core_product")
        db.execute("DELETE FROM core_batch")
        db.commit()

        reload_taxonomy()
        inserted = 0
        errors = []
        for i, row in enumerate(reader):
            try:
                pid = i + 1
                brand = row.get("brand", "")
                if not brand:
                    pname = row.get("product_name", row.get("title", ""))
                    if pname and " by " in pname:
                        brand = pname.rsplit(" by ", 1)[1].strip()
                ptype = row.get("product_type", "")
                if not ptype:
                    ptype = row.get("source_subcategory", "")
                db.execute("INSERT OR IGNORE INTO core_product VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (pid, row.get("product_number", row.get("SKU", "")),
                     row.get("product_name", row.get("title", "")),
                     row.get("product_description", row.get("description", "")),
                     row.get("image_url", row.get("image", "")), "[]",
                     row.get("source_category", row.get("category", "")),
                     row.get("source_subcategory", row.get("subcategory", "")),
                     row.get("collection_name", row.get("collection", "")),
                     row.get("materials", row.get("material", "")),
                     row.get("product_dimensions", row.get("dimensions", "")),
                     float(row.get("product_weight", row.get("weight", 0)) or 0) or None,
                     row.get("country_of_origin", row.get("country", "")),
                     row.get("product_url", row.get("url", "")),
                     row.get("model_number", ""), row.get("bullets", ""),
                     brand, ptype))
                text = f"{row.get('product_name','')} {row.get('product_description','')} {row.get('bullets','')} {row.get('source_category','')} {row.get('source_subcategory','')} {row.get('materials','')} {brand} {ptype}"
                node_id, confidence, alternatives, attrs = classify_product(
                    row.get("product_name", ""), row.get("source_category", ""),
                    row.get("source_subcategory", ""), row.get("product_description", ""),
                    row.get("materials", ""), brand)
                has_img = bool(row.get("image_url", row.get("image", "")))
                has_desc = bool(row.get("product_description", row.get("description", "")).strip())
                if confidence < 0.4:
                    status, reason, nr = "needs_review", "Low confidence", True
                elif not has_desc and not has_img:
                    status, reason, nr = "needs_review", "Missing description and image", True
                else:
                    status, reason, nr = "auto_classified", "", False
                db.execute("INSERT INTO core_classification VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (pid, confidence, status, json.dumps(alternatives), json.dumps(attrs),
                     nr, reason, datetime.now().isoformat(), None, "", pid, node_id))
                inserted += 1
            except Exception as e:
                errors.append(f"Row {i+1}: {str(e)}")
        db.commit()

        # Record batch
        try:
            now_iso = datetime.now().isoformat()
            total_rows = inserted + len(errors)
            db.execute("INSERT INTO core_batch (name, status, total_products, processed, succeeded, failed, started_at, finished_at) VALUES (?,?,?,?,?,?,?,?)",
                      (f"API Import: {file.filename}", "completed", total_rows, inserted, inserted, len(errors), now_iso, now_iso))
            db.commit()
        except Exception:
            pass

        db.close()
        return jsonify({"imported": inserted, "errors": len(errors), "error_details": errors[:10]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)), debug=True)
