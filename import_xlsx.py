import json
import os
import csv
import hashlib
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Shopify Taxonomy Nodes (69 nodes) with per-node attribute definitions
# ---------------------------------------------------------------------------
TAXONOMY_NODES = [
    {"id": 1, "shopify_id": "tax_0001", "name": "Living Room", "level": 0, "keywords": "sofa,couch,armchair,loveseat,sectional,ottoman,chaise,recliner,settee,futon", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 2, "shopify_id": "tax_0002", "name": "Sofas", "level": 1, "keywords": "sofa,couch,sectional,chesterfield,futon,daybed", "product_type_hint": "Sofas and Armchairs", "parent_id": 1, "attributes": ["Color", "Material", "Size", "Frame Material", "Cushion Type", "Seating Capacity", "Style"]},
    {"id": 3, "shopify_id": "tax_0003", "name": "Armchairs", "level": 1, "keywords": "armchair,accent chair,club chair,wingback,papasan,glider", "product_type_hint": "Sofas and Armchairs", "parent_id": 1, "attributes": ["Color", "Material", "Frame Material", "Style", "Swivel"]},
    {"id": 4, "shopify_id": "tax_0004", "name": "Coffee Tables", "level": 1, "keywords": "coffee table,cocktail table,center table", "product_type_hint": "Coffee Tables", "parent_id": 1, "attributes": ["Color", "Material", "Shape", "Storage", "Style"]},
    {"id": 5, "shopify_id": "tax_0005", "name": "TV Stands", "level": 1, "keywords": "tv stand,media console,entertainment center,television stand", "product_type_hint": "TV Stands", "parent_id": 1, "attributes": ["Color", "Material", "Fits TV Size", "Number of Shelves", "Style"]},
    {"id": 6, "shopify_id": "tax_0006", "name": "Bookshelves", "level": 1, "keywords": "bookshelf,bookcase,etagere,shelving unit", "product_type_hint": "Bookshelves", "parent_id": 1, "attributes": ["Color", "Material", "Number of Shelves", "Height", "Style"]},
    {"id": 7, "shopify_id": "tax_0007", "name": "Ottomans", "level": 1, "keywords": "ottoman,pouf,footrest,hassock", "product_type_hint": "Ottomans", "parent_id": 1, "attributes": ["Color", "Material", "Storage", "Shape", "Style"]},
    {"id": 8, "shopify_id": "tax_0008", "name": "Bedroom", "level": 0, "keywords": "bed,bedroom,mattress,headboard,nightstand,wardrobe,armoire,dresser", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 9, "shopify_id": "tax_0009", "name": "Beds", "level": 1, "keywords": "bed,bed frame,platform bed,bunk bed,canopy bed,poster bed,headboard", "product_type_hint": "Beds", "parent_id": 8, "attributes": ["Color", "Material", "Size", "Frame Material", "Style", "Headboard Included"]},
    {"id": 10, "shopify_id": "tax_0010", "name": "Dressers", "level": 1, "keywords": "dresser,chest of drawers,bureau,commode", "product_type_hint": "Dressers", "parent_id": 8, "attributes": ["Color", "Material", "Number of Drawers", "Style"]},
    {"id": 11, "shopify_id": "tax_0011", "name": "Nightstands", "level": 1, "keywords": "nightstand,bedside table,end table", "product_type_hint": "Nightstands", "parent_id": 8, "attributes": ["Color", "Material", "Number of Drawers", "Style"]},
    {"id": 12, "shopify_id": "tax_0012", "name": "Wardrobes", "level": 1, "keywords": "wardrobe,armoire,closet,closet system", "product_type_hint": "Wardrobes", "parent_id": 8, "attributes": ["Color", "Material", "Number of Doors", "Number of Shelves", "Style"]},
    {"id": 13, "shopify_id": "tax_0013", "name": "Office", "level": 0, "keywords": "office,desk,chair,filing cabinet,bookcase,workspace", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 14, "shopify_id": "tax_0014", "name": "Desks", "level": 1, "keywords": "desk,workstation,writing desk,standing desk,computer desk", "product_type_hint": "Desks", "parent_id": 13, "attributes": ["Color", "Material", "Shape", "Has Drawers", "Style"]},
    {"id": 15, "shopify_id": "tax_0015", "name": "Office Chairs", "level": 1, "keywords": "office chair,desk chair,task chair,executive chair,ergonomic chair", "product_type_hint": "Office Chairs", "parent_id": 13, "attributes": ["Color", "Material", "Armrests", "Adjustable Height", "Lumbar Support", "Style"]},
    {"id": 16, "shopify_id": "tax_0016", "name": "Filing Cabinets", "level": 1, "keywords": "filing cabinet,file cabinet,drawer unit", "product_type_hint": "Filing Cabinets", "parent_id": 13, "attributes": ["Color", "Material", "Number of Drawers", "Lockable", "Style"]},
    {"id": 17, "shopify_id": "tax_0017", "name": "Dining", "level": 0, "keywords": "dining,table,chairs,stool,bench,sideboard,bar", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 18, "shopify_id": "tax_0018", "name": "Dining Tables", "level": 1, "keywords": "dining table,kitchen table,conference table", "product_type_hint": "Dining Tables", "parent_id": 17, "attributes": ["Color", "Material", "Shape", "Seating Capacity", "Extension", "Style"]},
    {"id": 19, "shopify_id": "tax_0019", "name": "Dining Chairs", "level": 1, "keywords": "dining chair,kitchen chair,side chair", "product_type_hint": "Dining Chairs", "parent_id": 17, "attributes": ["Color", "Material", "Upholstered", "Style"]},
    {"id": 20, "shopify_id": "tax_0020", "name": "Bar Stools", "level": 1, "keywords": "bar stool,counter stool,backless stool,swivel stool", "product_type_hint": "Bar Stools", "parent_id": 17, "attributes": ["Color", "Material", "Height", "Swivel", "Backrest", "Style"]},
    {"id": 21, "shopify_id": "tax_0021", "name": "Buffets & Sideboards", "level": 1, "keywords": "buffet,sideboard,credenza,serving cabinet", "product_type_hint": "Buffets and Sideboards", "parent_id": 17, "attributes": ["Color", "Material", "Number of Doors", "Number of Drawers", "Style"]},
    {"id": 22, "shopify_id": "tax_0022", "name": "Dining Sets", "level": 1, "keywords": "dining set,kitchen set,dinette", "product_type_hint": "Dining Sets", "parent_id": 17, "attributes": ["Color", "Material", "Seating Capacity", "Table Shape", "Style"]},
    {"id": 23, "shopify_id": "tax_0023", "name": "Outdoor", "level": 0, "keywords": "outdoor,patio,garden,deck,backyard,terrace", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 24, "shopify_id": "tax_0024", "name": "Patio Furniture", "level": 1, "keywords": "patio,terrace,deck,outdoor furniture", "product_type_hint": "Patio Furniture", "parent_id": 23, "attributes": ["Color", "Material", "Weather Resistant", "Style"]},
    {"id": 25, "shopify_id": "tax_0025", "name": "Outdoor Dining", "level": 1, "keywords": "outdoor dining,patio table,patio chair,patio set", "product_type_hint": "Outdoor Dining", "parent_id": 23, "attributes": ["Color", "Material", "Seating Capacity", "Weather Resistant", "Style"]},
    {"id": 26, "shopify_id": "tax_0026", "name": "Outdoor Seating", "level": 1, "keywords": "outdoor sofa,adirondack,hammock,bench,loveseat outdoor", "product_type_hint": "Outdoor Seating", "parent_id": 23, "attributes": ["Color", "Material", "Weather Resistant", "Style"]},
    {"id": 27, "shopify_id": "tax_0027", "name": "Bath", "level": 0, "keywords": "bath,bathroom,vanity,mirror,cabinet,shelf,towel", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 28, "shopify_id": "tax_0028", "name": "Bathroom Vanities", "level": 1, "keywords": "bathroom vanity,vanity cabinet,sink basin,vanity sink", "product_type_hint": "Bathroom Vanities", "parent_id": 27, "attributes": ["Color", "Material", "Number of Drawers", "Number of Doors", "Sink Included", "Style"]},
    {"id": 29, "shopify_id": "tax_0029", "name": "Bathroom Storage", "level": 1, "keywords": "bathroom cabinet,medicine cabinet,bathroom shelf,bathroom storage", "product_type_hint": "Bathroom Storage", "parent_id": 27, "attributes": ["Color", "Material", "Number of Shelves", "Mount Type", "Style"]},
    {"id": 30, "shopify_id": "tax_0030", "name": "Bathroom Mirrors", "level": 1, "keywords": "bathroom mirror,vanity mirror,wall mirror", "product_type_hint": "Bathroom Mirrors", "parent_id": 27, "attributes": ["Color", "Shape", "Frame Material", "Mount Type", "Lighted", "Style"]},
    {"id": 31, "shopify_id": "tax_0031", "name": "Decor", "level": 0, "keywords": "decor,decoration,accent,ornament,figurine,vase,candle", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 32, "shopify_id": "tax_0032", "name": "Wall Art", "level": 1, "keywords": "wall art,painting,print,poster,canvas,frame", "product_type_hint": "Wall Art", "parent_id": 31, "attributes": ["Color", "Material", "Size", "Frame Included", "Style"]},
    {"id": 33, "shopify_id": "tax_0033", "name": "Mirrors", "level": 1, "keywords": "mirror,wall mirror,floor mirror,decorative mirror", "product_type_hint": "Mirrors", "parent_id": 31, "attributes": ["Color", "Shape", "Frame Material", "Mount Type", "Style"]},
    {"id": 34, "shopify_id": "tax_0034", "name": "Vases", "level": 1, "keywords": "vase,flower pot,planter,jar", "product_type_hint": "Vases", "parent_id": 31, "attributes": ["Color", "Material", "Shape", "Size", "Style"]},
    {"id": 35, "shopify_id": "tax_0035", "name": "Candles & Holders", "level": 1, "keywords": "candle,candle holder,candleholder,lantern", "product_type_hint": "Candles and Holders", "parent_id": 31, "attributes": ["Color", "Material", "Scented", "Size", "Style"]},
    {"id": 36, "shopify_id": "tax_0036", "name": "Clocks", "level": 1, "keywords": "clock,wall clock,desk clock,alarm clock", "product_type_hint": "Clocks", "parent_id": 31, "attributes": ["Color", "Material", "Display Type", "Style"]},
    {"id": 37, "shopify_id": "tax_0037", "name": "Lighting", "level": 0, "keywords": "light,lighting,lamp,fixture,chandelier,sconce,pendant", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 38, "shopify_id": "tax_0038", "name": "Table Lamps", "level": 1, "keywords": "table lamp,desk lamp,bedside lamp", "product_type_hint": "Table Lamps", "parent_id": 37, "attributes": ["Color", "Material", "Bulb Type", "Shade Material", "Style"]},
    {"id": 39, "shopify_id": "tax_0039", "name": "Floor Lamps", "level": 1, "keywords": "floor lamp,standing lamp,torchiere", "product_type_hint": "Floor Lamps", "parent_id": 37, "attributes": ["Color", "Material", "Bulb Type", "Adjustable", "Style"]},
    {"id": 40, "shopify_id": "tax_0040", "name": "Chandeliers", "level": 1, "keywords": "chandelier,pendant light,ceiling light", "product_type_hint": "Chandeliers", "parent_id": 37, "attributes": ["Color", "Material", "Number of Lights", "Style"]},
    {"id": 41, "shopify_id": "tax_0041", "name": "Wall Sconces", "level": 1, "keywords": "wall sconce,sconce,wall light", "product_type_hint": "Wall Sconces", "parent_id": 37, "attributes": ["Color", "Material", "Style"]},
    {"id": 42, "shopify_id": "tax_0042", "name": "String Lights", "level": 1, "keywords": "string lights,fairy lights,led strip,rope light", "product_type_hint": "String Lights", "parent_id": 37, "attributes": ["Color", "Length", "Bulb Type", "Indoor/Outdoor", "Style"]},
    {"id": 43, "shopify_id": "tax_0043", "name": "Rugs", "level": 0, "keywords": "rug,carpet,runner,mat,doormat", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 44, "shopify_id": "tax_0044", "name": "Area Rugs", "level": 1, "keywords": "area rug,rug,carpet,shag rug", "product_type_hint": "Area Rugs", "parent_id": 43, "attributes": ["Color", "Material", "Size", "Shape", "Pile Height", "Style"]},
    {"id": 45, "shopify_id": "tax_0045", "name": "Runners", "level": 1, "keywords": "runner,hallway runner,kitchen runner", "product_type_hint": "Runners", "parent_id": 43, "attributes": ["Color", "Material", "Size", "Style"]},
    {"id": 46, "shopify_id": "tax_0046", "name": "Doormats", "level": 1, "keywords": "doormat,door mat,welcome mat", "product_type_hint": "Doormats", "parent_id": 43, "attributes": ["Color", "Material", "Size", "Style"]},
    {"id": 47, "shopify_id": "tax_0047", "name": "Pillows & Throws", "level": 0, "keywords": "pillow,throw,blanket,cushion,quilt,duvet", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 48, "shopify_id": "tax_0048", "name": "Throw Pillows", "level": 1, "keywords": "throw pillow,accent pillow,decorative pillow,cushion", "product_type_hint": "Throw Pillows", "parent_id": 47, "attributes": ["Color", "Material", "Size", "Shape", "Fill Type", "Style"]},
    {"id": 49, "shopify_id": "tax_0049", "name": "Throw Blankets", "level": 1, "keywords": "throw blanket,throw,blanket,coverlet", "product_type_hint": "Throw Blankets", "parent_id": 47, "attributes": ["Color", "Material", "Size", "Weight", "Style"]},
    {"id": 50, "shopify_id": "tax_0050", "name": "Curtains", "level": 0, "keywords": "curtain,drapes,blinds,shades,window treatment", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 51, "shopify_id": "tax_0051", "name": "Curtains", "level": 1, "keywords": "curtain,drapes,panel,window treatment", "product_type_hint": "Curtains", "parent_id": 50, "attributes": ["Color", "Material", "Size", "Opacity", "Rod Type", "Style"]},
    {"id": 52, "shopify_id": "tax_0052", "name": "Blinds", "level": 1, "keywords": "blinds,shades,venetian blinds,roller blinds", "product_type_hint": "Blinds", "parent_id": 50, "attributes": ["Color", "Material", "Size", "Operation Type", "Style"]},
    {"id": 53, "shopify_id": "tax_0053", "name": "Storage & Organization", "level": 0, "keywords": "storage,organizer,shelf,rack,bin,basket,box", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 54, "shopify_id": "tax_0054", "name": "Baskets", "level": 1, "keywords": "basket,bin,container,tote", "product_type_hint": "Baskets", "parent_id": 53, "attributes": ["Color", "Material", "Size", "Style"]},
    {"id": 55, "shopify_id": "tax_0055", "name": "Shoe Racks", "level": 1, "keywords": "shoe rack,shoe storage,boot rack", "product_type_hint": "Shoe Racks", "parent_id": 53, "attributes": ["Color", "Material", "Number of Tiers", "Style"]},
    {"id": 56, "shopify_id": "tax_0056", "name": "Coat Racks", "level": 1, "keywords": "coat rack,coat hook,hall tree,wall hooks", "product_type_hint": "Coat Racks", "parent_id": 53, "attributes": ["Color", "Material", "Number of Hooks", "Style"]},
    {"id": 57, "shopify_id": "tax_0057", "name": "Shelving", "level": 1, "keywords": "shelf,shelving,storage rack,wire rack", "product_type_hint": "Shelving", "parent_id": 53, "attributes": ["Color", "Material", "Number of Shelves", "Weight Capacity", "Style"]},
    {"id": 58, "shopify_id": "tax_0058", "name": "Kitchen & Dining Storage", "level": 1, "keywords": "kitchen storage,spice rack,pantry,food storage", "product_type_hint": "Kitchen and Dining Storage", "parent_id": 53, "attributes": ["Color", "Material", "Size", "Style"]},
    {"id": 59, "shopify_id": "tax_0059", "name": "Garden", "level": 0, "keywords": "garden,plant,flower,pot,planter,tool,fence", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 60, "shopify_id": "tax_0060", "name": "Plants & Seeds", "level": 1, "keywords": "plant,seed,flower,herb,seedling", "product_type_hint": "Plants and Seeds", "parent_id": 59, "attributes": ["Plant Type", "Indoor/Outdoor", "Size"]},
    {"id": 61, "shopify_id": "tax_0061", "name": "Gardening Tools", "level": 1, "keywords": "garden tool,shovel,rake,pruner,hose", "product_type_hint": "Gardening Tools", "parent_id": 59, "attributes": ["Material", "Size", "Type"]},
    {"id": 62, "shopify_id": "tax_0062", "name": "Planters & Pots", "level": 1, "keywords": "planter,flower pot,hanging planter,window box", "product_type_hint": "Planters and Pots", "parent_id": 59, "attributes": ["Color", "Material", "Size", "Shape", "Drainage", "Style"]},
    {"id": 63, "shopify_id": "tax_0063", "name": "Outdoor Living", "level": 1, "keywords": "patio heater,fire pit,garden decor,umbrella", "product_type_hint": "Outdoor Living", "parent_id": 59, "attributes": ["Color", "Material", "Size", "Style"]},
    {"id": 64, "shopify_id": "tax_0064", "name": "Trash & Recycling", "level": 0, "keywords": "trash,garbage,recycle,waste,bin,can", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 65, "shopify_id": "tax_0065", "name": "Trash Cans", "level": 1, "keywords": "trash can,garbage can,waste basket,trash bin", "product_type_hint": "Trash Cans", "parent_id": 64, "attributes": ["Color", "Material", "Capacity", "Lid Type", "Style"]},
    {"id": 66, "shopify_id": "tax_0066", "name": "Recycling Bins", "level": 1, "keywords": "recycling bin,recycle bin", "product_type_hint": "Recycling Bins", "parent_id": 64, "attributes": ["Color", "Material", "Capacity", "Style"]},
    {"id": 67, "shopify_id": "tax_0067", "name": "Kids Furniture", "level": 0, "keywords": "kids,children,child,toddler,baby,nursery", "product_type_hint": "", "parent_id": None, "attributes": []},
    {"id": 68, "shopify_id": "tax_0068", "name": "Kids Beds", "level": 1, "keywords": "kids bed,child bed,toddler bed,bunk bed,loft bed", "product_type_hint": "Kids Beds", "parent_id": 67, "attributes": ["Color", "Material", "Size", "Style"]},
    {"id": 69, "shopify_id": "tax_0069", "name": "Kids Storage", "level": 1, "keywords": "kids storage,toy box,toy storage,bookshelf kids", "product_type_hint": "Kids Storage", "parent_id": 67, "attributes": ["Color", "Material", "Size", "Style"]},
]

NODE_MAP = {n["id"]: n for n in TAXONOMY_NODES}

# ---------------------------------------------------------------------------
# Classification mapping
# ---------------------------------------------------------------------------
CATEGORY_MAP = {
    ("Living Room", "Sofas and Armchairs"): 2,
    ("Living Room", "Coffee Tables"): 4,
    ("Living Room", "TV Stands"): 5,
    ("Living Room", "Bookshelves"): 6,
    ("Living Room", "Ottomans"): 7,
    ("Living Room", "Decor"): 31,
    ("Living Room", "Lighting"): 38,
    ("Bedroom", "Beds"): 9,
    ("Bedroom", "Dressers"): 10,
    ("Bedroom", "Nightstands"): 11,
    ("Bedroom", "Wardrobes"): 12,
    ("Office", "Desks"): 14,
    ("Office", "Chairs"): 15,
    ("Office", "Filing Cabinets"): 16,
    ("Dining", "Tables"): 18,
    ("Dining", "Chairs"): 19,
    ("Dining", "Bar Stools"): 20,
    ("Dining", "Buffets and Sideboards"): 21,
    ("Dining", "Dining Sets"): 22,
    ("Outdoor", "Patio Furniture"): 24,
    ("Outdoor", "Outdoor Dining"): 25,
    ("Outdoor", "Outdoor Seating"): 26,
    ("Bath", "Bathroom Vanities"): 28,
    ("Bath", "Bathroom Storage"): 29,
    ("Bath", "Bathroom Mirrors"): 30,
    ("Decor", "Wall Art"): 32,
    ("Decor", "Mirrors"): 33,
    ("Decor", "Vases"): 34,
    ("Decor", "Candles and Holders"): 35,
    ("Decor", "Clocks"): 36,
    ("Lighting", "Table Lamps"): 38,
    ("Lighting", "Floor Lamps"): 39,
    ("Lighting", "Chandeliers"): 40,
    ("Lighting", "Wall Sconces"): 41,
    ("Lighting", "String Lights"): 42,
    ("Rugs", "Area Rugs"): 44,
    ("Rugs", "Runners"): 45,
    ("Rugs", "Doormats"): 46,
    ("Pillows & Throws", "Throw Pillows"): 48,
    ("Pillows & Throws", "Throw Blankets"): 49,
    ("Curtains", "Curtains"): 51,
    ("Curtains", "Blinds"): 52,
    ("Storage & Organization", "Baskets"): 54,
    ("Storage & Organization", "Shoe Racks"): 55,
    ("Storage & Organization", "Coat Racks"): 56,
    ("Storage & Organization", "Shelving"): 57,
    ("Storage & Organization", "Kitchen and Dining Storage"): 58,
    ("Garden", "Plants and Seeds"): 60,
    ("Garden", "Gardening Tools"): 61,
    ("Garden", "Planters and Pots"): 62,
    ("Garden", "Outdoor Living"): 63,
    ("Trash & Recycling", "Trash Cans"): 65,
    ("Trash & Recycling", "Recycling Bins"): 66,
    ("Kids Furniture", "Kids Beds"): 68,
    ("Kids Furniture", "Kids Storage"): 69,
}

KEYWORD_CLASSIFICATION = {
    "sofa": 2, "couch": 2, "sectional": 2, "loveseat": 2,
    "armchair": 3, "accent chair": 3, "club chair": 3,
    "coffee table": 4, "cocktail table": 4,
    "tv stand": 5, "media console": 5, "entertainment center": 5,
    "bookshelf": 6, "bookcase": 6, "etagere": 6,
    "ottoman": 7, "pouf": 7, "footrest": 7,
    "bed": 9, "headboard": 9, "mattress": 9,
    "dresser": 10, "chest of drawers": 10,
    "nightstand": 11, "bedside table": 11,
    "wardrobe": 12, "armoire": 12,
    "desk": 14, "workstation": 14,
    "office chair": 15, "desk chair": 15,
    "filing cabinet": 16,
    "dining table": 18, "kitchen table": 18,
    "dining chair": 19, "kitchen chair": 19,
    "bar stool": 20, "counter stool": 20,
    "buffet": 21, "sideboard": 21,
    "dining set": 22,
    "patio": 24, "outdoor furniture": 24,
    "outdoor dining": 25, "patio table": 25,
    "outdoor sofa": 26, "adirondack": 26,
    "vanity": 28, "bathroom vanity": 28,
    "bathroom cabinet": 29, "medicine cabinet": 29,
    "bathroom mirror": 30,
    "wall art": 32, "painting": 32, "print": 32,
    "mirror": 33, "wall mirror": 33,
    "vase": 34, "flower pot": 34,
    "candle": 35, "candle holder": 35,
    "clock": 36, "wall clock": 36,
    "table lamp": 38, "desk lamp": 38,
    "floor lamp": 39,
    "chandelier": 40, "pendant light": 40,
    "wall sconce": 41,
    "string lights": 42,
    "area rug": 44, "rug": 44, "carpet": 44,
    "runner": 45,
    "doormat": 46,
    "throw pillow": 48, "accent pillow": 48,
    "throw blanket": 49, "blanket": 49,
    "curtain": 51, "drapes": 51,
    "blinds": 52,
    "basket": 54, "bin": 54,
    "shoe rack": 55,
    "coat rack": 56,
    "shelf": 57, "shelving": 57,
    "kitchen storage": 58,
    "plant": 60, "seed": 60,
    "garden tool": 61, "shovel": 61,
    "planter": 62,
    "patio heater": 63, "fire pit": 63,
    "trash can": 65, "garbage can": 65,
    "recycling bin": 66,
    "kids bed": 68, "toddler bed": 68,
    "toy box": 69, "toy storage": 69,
}

# ---------------------------------------------------------------------------
# Attribute detection helpers
# ---------------------------------------------------------------------------
COLORS = ["white","black","brown","gray","grey","beige","navy","blue","red","green","gold","silver","natural","walnut","oak","espresso","tan","cream","pink","purple","yellow","orange","charcoal","mahogany","teak","acacia"]
MATERIALS = {"wood":"Wood","metal":"Metal","steel":"Steel","iron":"Iron","aluminum":"Aluminum","fabric":"Fabric","leather":"Leather","velvet":"Velvet","linen":"Linen","cotton":"Cotton","polyester":"Polyester","plastic":"Plastic","marble":"Marble","glass":"Glass","rattan":"Rattan","wicker":"Wicker","bamboo":"Bamboo","mango wood":"Mango Wood","rubberwood":"Rubberwood","acacia":"Acacia","eucalyptus":"Eucalyptus","teak":"Teak","engineered wood":"Engineered Wood","mdf":"MDF","particleboard":"Particleboard","ceramic":"Ceramic","concrete":"Concrete","stone":"Stone","resin":"Resin","rope":"Rope","boucle":"Boucle","chenille":"Chenille","microfiber":"Microfiber","faux leather":"Faux Leather","bonded leather":"Bonded Leather","top grain leather":"Top Grain Leather","solid wood":"Solid Wood"}
SHAPES = {"round":"Round","square":"Square","rectangular":"Rectangular","oval":"Oval","heart":"Heart","hexagonal":"Hexagonal","c-shaped":"C-Shaped","half-moon":"Half Moon"}

def detect_color(text):
    t = text.lower()
    for c in COLORS:
        if c in t:
            return c.title()
    return ""

def detect_materials(text):
    t = text.lower()
    found = []
    for kw, val in MATERIALS.items():
        if kw in t and val not in found:
            found.append(val)
    return found[:3]

def detect_shape(text):
    t = text.lower()
    for kw, val in SHAPES.items():
        if kw in t:
            return val
    return ""

def detect_size(text):
    t = text.lower()
    if any(w in t for w in ["king","queen","full","twin","california king"]):
        for s in ["California King","King","Queen","Full","Twin"]:
            if s.lower() in t:
                return s
    if any(w in t for w in ["large","big"]):
        return "Large"
    if "small" in t:
        return "Small"
    if "medium" in t:
        return "Medium"
    return ""

def detect_style(text):
    t = text.lower()
    styles = {
        "modern":"Modern","contemporary":"Contemporary","mid-century":"Mid-Century",
        "farmhouse":"Farmhouse","industrial":"Industrial","scandinavian":"Scandinavian",
        "bohemian":"Bohemian","traditional":"Traditional","rustic":"Rustic",
        "minimalist":"Minimalist","transitional":"Transitional","coastal":"Coastal",
        "art deco":"Art Deco","vintage":"Vintage","retro":"Retro","glam":"Glam",
        "victorian":"Victorian","colonial":"Colonial","japanese":"Japanese",
        "asian":"Asian","moroccan":"Moroccan","mediterranean":"Mediterranean",
        "loft":"Loft","nordic":"Nordic","zen":"Zen","baroque":"Baroque",
    }
    for kw, val in styles.items():
        if kw in t:
            return val
    return ""

def detect_weight_capacity(text):
    import re
    m = re.search(r'(?:weight\s*capacity|max(?:imum)?\s*(?:weight|capacity))\s*[:\-]?\s*(\d+)', text, re.I)
    return int(m.group(1)) if m else None

def detect_assembly(text):
    t = text.lower()
    if "assembly required" in t or "some assembly" in t or "easy assembly" in t:
        return True
    if "no assembly" in t or "fully assembled" in t or "comes assembled" in t:
        return False
    return None

def extract_brand(product_name, collection_name):
    if product_name and " by " in product_name:
        parts = product_name.rsplit(" by ", 1)
        if len(parts) == 2 and len(parts[1]) < 40:
            return parts[1].strip()
    if collection_name and collection_name.lower() not in ("", "n/a", "none"):
        return collection_name.strip()
    return ""

def extract_product_type(product_name, source_subcategory):
    if source_subcategory and source_subcategory.lower() not in ("", "n/a", "none"):
        return source_subcategory.strip()
    if product_name:
        words = product_name.split()
        for kw in ["sofa","couch","chair","table","desk","bed","lamp","mirror","shelf","bookcase","dresser","nightstand","wardrobe","stool","bench","ottoman","rug","curtain","blinds","vase","clock","basket","planter","trash can"]:
            if kw in product_name.lower():
                return kw.title()
    return ""

# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------
def classify_product(product_name, source_category, source_subcategory, description="", materials=""):
    name_lower = (product_name or "").lower()
    desc_lower = (description or "").lower()
    mat_lower = (materials or "").lower()
    combined = f"{name_lower} {desc_lower} {mat_lower}"

    scores = []
    # 1) Exact category map
    key = (source_category, source_subcategory)
    if key in CATEGORY_MAP:
        nid = CATEGORY_MAP[key]
        scores.append((nid, 0.95))

    # 2) Keyword scoring per node
    for kw, nid in KEYWORD_CLASSIFICATION.items():
        if kw in combined:
            node = NODE_MAP.get(nid, {})
            kws = [k.strip() for k in node.get("keywords", "").split(",") if k.strip()]
            matched = sum(1 for k in kws if k in combined)
            score = min(0.6 + 0.1 * matched, 0.94)
            scores.append((nid, score))

    # 3) Parent keyword boost
    if scores:
        top_nid = max(scores, key=lambda x: x[1])[0]
        parent_id = NODE_MAP.get(top_nid, {}).get("parent_id")
        if parent_id:
            parent = NODE_MAP.get(parent_id, {})
            pkws = [k.strip() for k in parent.get("keywords", "").split(",") if k.strip()]
            parent_matched = sum(1 for k in pkws if k in combined)
            if parent_matched > 0:
                boost = min(0.05 * parent_matched, 0.1)
                scores = [(nid, min(s + boost, 0.99)) if nid == top_nid else (nid, s) for nid, s in scores]

    if not scores:
        return 1, 0.30, []

    scores.sort(key=lambda x: -x[1])
    best_nid, best_score = scores[0]

    # If best score is very low, flag for review
    if best_score < 0.4:
        best_nid = scores[0][0]
        best_score = max(best_score, 0.25)

    # Top 3 alternatives
    alternatives = []
    seen = {best_nid}
    for nid, sc in scores[1:]:
        if nid not in seen and len(alternatives) < 3:
            node = NODE_MAP.get(nid, {})
            alternatives.append({
                "name": node.get("name", ""),
                "confidence": round(sc, 3),
                "path": _build_path(nid),
            })
            seen.add(nid)

    return best_nid, round(best_score, 3), alternatives

def _build_path(node_id):
    parts = []
    nid = node_id
    while nid:
        n = NODE_MAP.get(nid)
        if not n:
            break
        parts.append(n["name"])
        nid = n.get("parent_id")
    return " > ".join(reversed(parts))

def detect_attributes(text, node_id):
    node = NODE_MAP.get(node_id, {})
    attrs = {}
    text_combined = text

    color = detect_color(text_combined)
    if color:
        attrs["Color"] = color

    mats = detect_materials(text_combined)
    if mats:
        attrs["Material"] = ", ".join(mats)

    shape = detect_shape(text_combined)
    if shape:
        attrs["Shape"] = shape

    size = detect_size(text_combined)
    if size:
        attrs["Size"] = size

    style = detect_style(text_combined)
    if style:
        attrs["Style"] = style

    wc = detect_weight_capacity(text_combined)
    if wc:
        attrs["Weight Capacity"] = f"{wc} lbs"

    assembly = detect_assembly(text_combined)
    if assembly is not None:
        attrs["Assembly Required"] = "Yes" if assembly else "No"

    return attrs

# ---------------------------------------------------------------------------
# Product reader
# ---------------------------------------------------------------------------
def read_xlsx(file_path):
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    products = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if row[0] is None:
                continue
            images = []
            for i in range(20, 39):
                if i < len(row) and row[i]:
                    images.append(str(row[i]))
            product = {
                "product_number": str(row[0] or ""),
                "model_number": str(row[1] or ""),
                "source_category": str(row[2] or ""),
                "source_subcategory": str(row[3] or ""),
                "collection_name": str(row[4] or ""),
                "product_name": str(row[7] or ""),
                "product_description": str(row[8] or ""),
                "bullets": str(row[9] or ""),
                "materials": str(row[12] or ""),
                "product_dimensions": str(row[13] or ""),
                "product_weight": float(row[11]) if row[11] else None,
                "image_url": str(row[21]) if row[21] else "",
                "all_image_urls": json.dumps(images),
                "product_url": str(row[47]) if len(row) > 47 and row[47] else "",
                "cost": float(row[18]) if row[18] else None,
                "msrp": float(row[20]) if row[20] else None,
                "country_of_origin": str(row[17] or ""),
                "brand": "",
                "product_type": "",
            }
            product["brand"] = extract_brand(product["product_name"], product["collection_name"])
            product["product_type"] = extract_product_type(product["product_name"], product["source_subcategory"])
            products.append(product)
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})
    wb.close()
    return products, errors

def read_csv(file_path):
    products = []
    errors = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, start=2):
            try:
                product = {
                    "product_number": row.get("product_number", row.get("SKU", "")),
                    "model_number": row.get("model_number", ""),
                    "source_category": row.get("source_category", row.get("category", "")),
                    "source_subcategory": row.get("source_subcategory", row.get("subcategory", "")),
                    "collection_name": row.get("collection_name", row.get("collection", "")),
                    "product_name": row.get("product_name", row.get("title", "")),
                    "product_description": row.get("product_description", row.get("description", "")),
                    "bullets": row.get("bullets", ""),
                    "materials": row.get("materials", row.get("material", "")),
                    "product_dimensions": row.get("product_dimensions", row.get("dimensions", "")),
                    "product_weight": float(row.get("product_weight", row.get("weight", 0)) or 0) or None,
                    "image_url": row.get("image_url", row.get("image", "")),
                    "all_image_urls": row.get("all_image_urls", "[]"),
                    "product_url": row.get("product_url", row.get("url", "")),
                    "cost": float(row.get("cost", 0) or 0) or None,
                    "msrp": float(row.get("msrp", 0) or 0) or None,
                    "country_of_origin": row.get("country_of_origin", row.get("country", "")),
                    "brand": row.get("brand", ""),
                    "product_type": row.get("product_type", ""),
                }
                if not product["brand"]:
                    product["brand"] = extract_brand(product["product_name"], product["collection_name"])
                if not product["product_type"]:
                    product["product_type"] = extract_product_type(product["product_name"], product["source_subcategory"])
                products.append(product)
            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})
    return products, errors

# ---------------------------------------------------------------------------
# Seed chunk creation with checkpoint/resume
# ---------------------------------------------------------------------------
def create_seed_chunks(products, output_dir, chunk_size=250, checkpoint_file=None):
    os.makedirs(output_dir, exist_ok=True)

    # Checkpoint for resume
    start_idx = 0
    if checkpoint_file and os.path.exists(checkpoint_file):
        with open(checkpoint_file) as f:
            cp = json.load(f)
            start_idx = cp.get("last_index", 0)
        print(f"Resuming from product {start_idx}")

    total = len(products)
    chunks_created = 0
    for chunk_start in range(start_idx, total, chunk_size):
        chunk_end = min(chunk_start + chunk_size, total)
        chunk = products[chunk_start:chunk_end]
        chunk_idx = chunk_start // chunk_size

        seed_data = {"products": [], "taxonomy_nodes": TAXONOMY_NODES, "classifications": []}

        for prod_idx, product in enumerate(chunk):
            product_id = chunk_start + prod_idx + 1
            product["id"] = product_id
            product["imported_at"] = datetime.now().isoformat()

            text = f"{product['product_name']} {product['product_description']} {product['bullets']} {product['source_category']} {product['source_subcategory']} {product['collection_name']} {product['materials']} {product.get('brand','')} {product.get('product_type','')}"
            node_id, confidence, alternatives = classify_product(
                product["product_name"], product["source_category"],
                product["source_subcategory"], product["product_description"],
                product["materials"]
            )

            attrs = detect_attributes(text, node_id)

            has_image = bool(product.get("image_url"))
            has_desc = bool(product.get("product_description", "").strip())
            missing_data = not has_desc and not has_image

            if confidence < 0.4:
                status = "needs_review"
                reason = "Low confidence classification"
                needs_review = True
            elif missing_data:
                status = "needs_review"
                reason = "Missing description and image"
                needs_review = True
            elif not has_desc:
                status = "needs_review"
                reason = "Missing product description"
                needs_review = True
            else:
                status = "auto_classified"
                reason = ""
                needs_review = False

            classification = {
                "id": product_id,
                "confidence": confidence,
                "status": status,
                "alternatives": json.dumps(alternatives),
                "detected_attributes": json.dumps(attrs),
                "needs_review": needs_review,
                "review_reason": reason,
                "classified_at": datetime.now().isoformat(),
                "reviewed_at": None,
                "reviewer_notes": "",
                "product_id": product_id,
                "taxonomy_node_id": node_id,
            }

            seed_data["products"].append(product)
            seed_data["classifications"].append(classification)

        chunk_file = os.path.join(output_dir, f"seed_chunk_{chunk_idx}.json")
        with open(chunk_file, "w") as f:
            json.dump(seed_data, f)
        chunks_created += 1
        print(f"  chunk_{chunk_idx}: products {chunk_start+1}-{chunk_end}")

        # Save checkpoint
        if checkpoint_file:
            with open(checkpoint_file, "w") as f:
                json.dump({"last_index": chunk_end, "total": total, "timestamp": datetime.now().isoformat()}, f)

    print(f"Done: {chunks_created} chunks, {total} products")
    return total

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python import_xlsx.py <product_list.xlsx|csv> [output_dir] [chunk_size]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "."
    chunk_size = int(sys.argv[3]) if len(sys.argv) > 3 else 250
    checkpoint = os.path.join(output_dir, ".import_checkpoint.json")

    print(f"Reading {input_file}...")
    if input_file.endswith(".xlsx"):
        products, errors = read_xlsx(input_file)
    elif input_file.endswith(".csv"):
        products, errors = read_csv(input_file)
    else:
        print("Unsupported file format. Use .xlsx or .csv")
        sys.exit(1)

    print(f"Read {len(products)} products ({len(errors)} errors)")
    if errors:
        print(f"First 5 errors: {errors[:5]}")

    print(f"Creating seed chunks (size={chunk_size})...")
    total = create_seed_chunks(products, output_dir, chunk_size, checkpoint)
    print(f"\nComplete: {total} products exported to {output_dir}")
